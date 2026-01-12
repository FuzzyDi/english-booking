import os
import psycopg2
from psycopg2.extras import RealDictCursor
from flask import Flask, render_template, request, redirect, url_for, session, make_response
from datetime import datetime, timedelta, date

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-in-production'

SLOTS = [
    "14:00-14:30", "14:30-15:00", "15:00-15:30", "15:30-16:00",
    "16:00-16:30", "16:30-17:00", "17:00-17:30", "17:30-18:00",
    "18:00-18:30", "18:30-19:00", "19:00-19:30", "19:30-20:00"
]

def get_db_connection():
    db_url = os.environ.get("DATABASE_URL")
    if not db_url:
        raise Exception("DATABASE_URL environment variable is required")
    return psycopg2.connect(db_url, cursor_factory=RealDictCursor)

def init_db():
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Таблица бронирований
            cur.execute("""
                CREATE TABLE IF NOT EXISTS bookings (
                    id SERIAL PRIMARY KEY,
                    name TEXT NOT NULL,
                    phone TEXT NOT NULL,
                    date DATE NOT NULL,
                    time_slot TEXT NOT NULL,
                    status TEXT NOT NULL DEFAULT 'confirmed',
                    attended INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # Таблица исключений расписания
            cur.execute("""
                CREATE TABLE IF NOT EXISTS availability_override (
                    id SERIAL PRIMARY KEY,
                    date DATE NOT NULL,
                    time_slot TEXT
                )
            """)

            # Проверка наличия столбца 'attended' (для обновлений)
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_name='bookings' AND column_name='attended'
            """)
            if cur.fetchone() is None:
                cur.execute("ALTER TABLE bookings ADD COLUMN attended INTEGER")

            conn.commit()
    finally:
        conn.close()

def mask_phone(phone):
    if len(phone) >= 7:
        return phone[:4] + '*' * (len(phone) - 7) + phone[-3:]
    return phone

def get_current_week_dates():
    today = date.today()
    if today.weekday() == 6:
        start = today + timedelta(days=1)
    else:
        start = today - timedelta(days=today.weekday())
    return [start + timedelta(days=i) for i in range(6)]

def cleanup_old_bookings():
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            today = date.today()
            last_sunday = today - timedelta(days=today.weekday() + 1)
            cur.execute("""
                DELETE FROM bookings 
                WHERE date <= %s AND status IN ('confirmed', 'cancelled')
            """, (last_sunday.isoformat(),))
            conn.commit()
    finally:
        conn.close()

def is_slot_available(target_date, time_slot):
    if target_date.weekday() >= 6:
        return False
    if time_slot not in SLOTS:
        return False

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Проверка: отключён ли день?
            cur.execute(
                "SELECT 1 FROM availability_override WHERE date = %s AND time_slot IS NULL",
                (target_date.isoformat(),)
            )
            if cur.fetchone():
                return False

            # Проверка: отключён ли слот?
            cur.execute(
                "SELECT 1 FROM availability_override WHERE date = %s AND time_slot = %s",
                (target_date.isoformat(), time_slot)
            )
            if cur.fetchone():
                return False

            # Занят ли слот (confirmed или pending)?
            cur.execute(
                """SELECT 1 FROM bookings 
                   WHERE date = %s AND time_slot = %s AND status IN ('confirmed', 'pending_cancellation')""",
                (target_date.isoformat(), time_slot)
            )
            if cur.fetchone():
                return False

        return True
    finally:
        conn.close()

def can_book_client(phone, target_date):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Один раз в день
            cur.execute(
                "SELECT 1 FROM bookings WHERE phone = %s AND date = %s AND status = 'confirmed'",
                (phone, target_date.isoformat())
            )
            if cur.fetchone():
                return False, "You are already booked for this day."

            # Макс. 3 в неделю
            monday = target_date - timedelta(days=target_date.weekday())
            sunday = monday + timedelta(days=6)
            cur.execute(
                """SELECT COUNT(*) FROM bookings 
                   WHERE phone = %s AND date BETWEEN %s AND %s AND status = 'confirmed'""",
                (phone, monday.isoformat(), sunday.isoformat())
            )
            count = cur.fetchone()['count']
            if count >= 3:
                return False, "Maximum 3 sessions per week."

        return True, ""
    finally:
        conn.close()

@app.route('/')
def index():
    now = datetime.now()
    if now.weekday() == 6 and now.hour >= 10:
        cleanup_old_bookings()

    week_dates = get_current_week_dates()
    days = []
    for d in week_dates:
        slots = []
        for slot in SLOTS:
            available = is_slot_available(d, slot)
            slots.append({'time': slot, 'available': available})
        days.append({
            'date': d,
            'formatted': d.strftime('%A, %b %d'),
            'slots': slots
        })
    return render_template('index.html', days=days)

@app.route('/book', methods=['POST'])
def book():
    name = request.form['name'].strip()
    phone = request.form['phone'].strip()
    date_str = request.form['date']
    time_slot = request.form['time_slot']

    if not name or not phone:
        return "Name and phone are required", 400

    try:
        target_date = datetime.fromisoformat(date_str).date()
    except:
        return "Invalid date", 400

    if not is_slot_available(target_date, time_slot):
        return "Slot is not available", 400

    can, msg = can_book_client(phone, target_date)
    if not can:
        return msg, 400

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """INSERT INTO bookings (name, phone, date, time_slot, status)
                   VALUES (%s, %s, %s, %s, 'confirmed')""",
                (name, phone, date_str, time_slot)
            )
            conn.commit()
    finally:
        conn.close()

    response = make_response(redirect(url_for('success')))
    response.set_cookie('user_phone', phone, max_age=7*24*60*60)
    return response

@app.route('/success')
def success():
    return """
    <div style="font-family: Poppins, sans-serif; background: #1a1a2e; color: white; min-height: 100vh; 
                display: flex; align-items: center; justify-content: center; text-align: center; padding: 20px;">
        <div>
            <h1 style="font-size: 2.5em; margin-bottom: 20px; color: #00f5d4;">Booked!</h1>
            <p>Your English session is confirmed.</p>
            <a href="/my-bookings" style="display: inline-block; margin-top: 20px; padding: 12px 24px; 
               background: #00f5d4; color: #0f172a; text-decoration: none; border-radius: 8px; font-weight: bold;">
                View My Bookings
            </a>
        </div>
    </div>
    """

# === User: My Bookings ===
@app.route('/my-bookings', methods=['GET', 'POST'])
def my_bookings():
    if request.method == 'POST':
        phone = request.form['phone'].strip()
    else:
        phone = request.cookies.get('user_phone')

    if phone:
        conn = get_db_connection()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """SELECT id, date, time_slot, status FROM bookings 
                       WHERE phone = %s AND date >= %s ORDER BY date""",
                    (phone, date.today().isoformat())
                )
                bookings = cur.fetchall()
            masked = mask_phone(phone)
            return render_template('my_bookings.html', bookings=bookings, phone=masked)
        finally:
            conn.close()
    
    return render_template('check_bookings.html')

@app.route('/cancel/<int:booking_id>', methods=['POST'])
def cancel_booking(booking_id):
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE bookings SET status = 'pending_cancellation' WHERE id = %s", (booking_id,))
            conn.commit()
    finally:
        conn.close()
    return redirect(url_for('my_bookings'))

# === Admin Panel ===
@app.route('/admin', methods=['GET', 'POST'])
def admin():
    if request.method == 'POST':
        if request.form.get('password') == 'adasiniqizi':
            session['admin'] = True
        else:
            return "Invalid password", 403

    if not session.get('admin'):
        return '''
        <form method="post" style="font-family: Poppins, sans-serif; max-width: 400px; margin: 100px auto; 
                padding: 20px; background: #16213e; border-radius: 10px; color: white;">
            <h2>Admin Login</h2>
            <input type="password" name="password" placeholder="Password" required 
                   style="width: 100%; padding: 10px; margin: 10px 0; border-radius: 5px; 
                          border: 1px solid #334155; background: #0f3460; color: white;">
            <button type="submit" style="width: 100%; padding: 10px; background: #00f5d4; 
                    color: #0f172a; border: none; border-radius: 5px; font-weight: bold;">Login</button>
        </form>
        '''

    today = date.today().isoformat()
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            # Сегодняшние записи
            cur.execute("""
                SELECT id, name, phone, time_slot, status, attended FROM bookings
                WHERE date = %s 
                ORDER BY time_slot
            """, (today,))
            today_bookings = cur.fetchall()

            # Все записи недели
            cur.execute("""
                SELECT id, name, phone, date, time_slot, status, attended FROM bookings
                WHERE date >= %s ORDER BY date, time_slot
            """, (date.today().isoformat(),))
            bookings = cur.fetchall()

            # Расписание
            dates = get_current_week_dates()
            schedule_data = []
            for d in dates:
                cur.execute(
                    "SELECT time_slot FROM availability_override WHERE date = %s",
                    (d.isoformat(),)
                )
                overrides = cur.fetchall()
                disabled_slots = set(row['time_slot'] for row in overrides if row['time_slot'] is not None)
                full_day_disabled = any(row['time_slot'] is None for row in overrides)
                schedule_data.append({
                    'date': d,
                    'full_disabled': full_day_disabled,
                    'disabled_slots': disabled_slots
                })

        masked_today = []
        for b in today_bookings:
            masked_today.append((
                b['id'], b['name'], mask_phone(b['phone']), b['time_slot'],
                b['status'], b['attended']
            ))

        masked_bookings = []
        for b in bookings:
            masked_bookings.append((
                b['id'], b['name'], mask_phone(b['phone']), b['date'],
                b['time_slot'], b['status'], b['attended']
            ))

        return render_template(
            'admin.html',
            today_bookings=masked_today,
            bookings=masked_bookings,
            schedule_data=schedule_data,
            slots=SLOTS,
            today=date.today().strftime('%A, %b %d')
        )
    finally:
        conn.close()

@app.route('/admin/update_schedule', methods=['POST'])
def update_schedule():
    if not session.get('admin'):
        return "Access denied", 403

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM availability_override")
            for key, value in request.form.items():
                if key.startswith('disable_'):
                    parts = key.replace('disable_', '').split('_')
                    if len(parts) == 1:
                        cur.execute("INSERT INTO availability_override (date, time_slot) VALUES (%s, NULL)", (parts[0],))
                    elif len(parts) == 2:
                        cur.execute("INSERT INTO availability_override (date, time_slot) VALUES (%s, %s)", (parts[0], parts[1]))
            conn.commit()
        return redirect(url_for('admin'))
    finally:
        conn.close()

@app.route('/admin/set_attendance/<int:booking_id>/<int:status>', methods=['POST'])
def set_attendance(booking_id, status):
    if not session.get('admin'):
        return "Access denied", 403
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE bookings SET attended = %s WHERE id = %s", (status, booking_id))
            conn.commit()
        return redirect(url_for('admin'))
    finally:
        conn.close()

@app.route('/admin/approve_cancel/<int:booking_id>', methods=['POST'])
def approve_cancel(booking_id):
    if not session.get('admin'):
        return "Access denied", 403
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE bookings SET status = 'cancelled' WHERE id = %s", (booking_id,))
            conn.commit()
        return redirect(url_for('admin'))
    finally:
        conn.close()

@app.route('/admin/reject_cancel/<int:booking_id>', methods=['POST'])
def reject_cancel(booking_id):
    if not session.get('admin'):
        return "Access denied", 403
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("UPDATE bookings SET status = 'confirmed' WHERE id = %s", (booking_id,))
            conn.commit()
        return redirect(url_for('admin'))
    finally:
        conn.close()

@app.route('/admin/export_excel')
def export_excel():
    if not session.get('admin'):
        return "Access denied", 403

    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT name, phone, date, time_slot, status,
                       CASE WHEN attended = 1 THEN 'Present'
                            WHEN attended = 0 THEN 'Absent'
                            ELSE 'Not marked' END as attendance
                FROM bookings
                WHERE date >= %s
                ORDER BY date, time_slot
            """, (date.today().isoformat(),))
            records = cur.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Bookings Report"

        headers = ["Name", "Phone", "Date", "Time Slot", "Status", "Attendance"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).font = header_font
            ws.cell(row=1, column=col).fill = header_fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

        for row in records:
            ws.append([row['name'], row['phone'], row['date'], row['time_slot'], row['status'], row['attendance']])

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='english_bookings_report.xlsx'
        )
    finally:
        conn.close()

@app.route('/admin/reports')
def admin_reports():
    if not session.get('admin'):
        return redirect(url_for('admin'))

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            today = date.today()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)

            total_slots = 6 * len(SLOTS)
            cur.execute("""
                SELECT COUNT(*) as cnt FROM bookings 
                WHERE date BETWEEN %s AND %s AND status IN ('confirmed', 'pending_cancellation')
            """, (week_start.isoformat(), week_end.isoformat()))
            booked = cur.fetchone()['cnt']

            cur.execute("SELECT COUNT(*) as cnt FROM bookings WHERE attended = 1")
            present = cur.fetchone()['cnt']

            cur.execute("SELECT COUNT(*) as cnt FROM bookings WHERE attended = 0")
            absent = cur.fetchone()['cnt']

            cur.execute("""
                SELECT name, phone, COUNT(*) as cnt 
                FROM bookings 
                GROUP BY phone, name
                ORDER BY cnt DESC 
                LIMIT 5
            """)
            top_students = cur.fetchall()

            load_by_day = []
            for i in range(6):
                d = week_start + timedelta(days=i)
                cur.execute("""
                    SELECT COUNT(*) as cnt FROM bookings 
                    WHERE date = %s AND status IN ('confirmed', 'pending_cancellation')
                """, (d.isoformat(),))
                cnt = cur.fetchone()['cnt']
                load_by_day.append({
                    'day': d.strftime('%A'),
                    'count': cnt,
                    'percent': round(cnt / len(SLOTS) * 100)
                })

        return render_template('admin_reports.html',
            total_slots=total_slots,
            booked=booked,
            load_percent=round(booked / total_slots * 100),
            present=present,
            absent=absent,
            top_students=top_students,
            load_by_day=load_by_day,
            week_start=week_start.strftime('%b %d'),
            week_end=week_end.strftime('%b %d')
        )
    finally:
        conn.close()

if __name__ == '__main__':
    init_db()
    port = int(os.environ.get("PORT", 5000))
    app.run(host='0.0.0.0', port=port)